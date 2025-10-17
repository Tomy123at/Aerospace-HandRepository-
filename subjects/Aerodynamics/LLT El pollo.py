# -*- coding: utf-8 -*-
"""
Created on Wed Sep  4 21:42:42 2024

@author: garci
"""

# El siguiente codigo se realizó para el calculo de las caracteristicas aerodinámicas de un ala finita
# por el método LLT.
import numpy as np
import pandas as pd
import os
import sys
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

import Tabla_ISA_module 

# Parametros necesarios para los cálculos:
# Geometría del ala.
bw = 3.37  # Embergadura del ala [m]
Croot = 0.958     # Cuerda de la raiz del ala [m]
Ctip = 0.958  # Cuerda de la punta del ala [m]
lambdaw = Ctip/Croot    # Taper ratio
Sw = (Croot + Ctip)*bw/2    # Área de la mitad del ala
AR = bw**2/Sw       # Acpect ratio
betaw = 0*np.pi/180      # Wing twist angle [deg]
htip = Ctip*np.sin(betaw)
AoA = np.radians(np.arange(-12, 12 + 1, 1))   # Vector  geometric Angle of attack [deg]
# Airfoil aerodynamics characteristics:
# Airfoil reference: Root: NACA 0015 Tip: NACA 0012
a0root = 6.54317802    # Root Lift slope [1/rad]
a0tip = 6.54317802     # Tip Lift slope [1/rad]
aL_0_root =  0*np.pi/180 # Root Anfle of attack for L = 0 [rad]
aL_0_tip = 0*np.pi/180   # Tip Anfle of attack for L = 0 [rad]
#Cd_0 = np.arange()
# Flying (airflow) characteristics:
T,P,rho,avel = Tabla_ISA_module.Temp_Presion_Densidad(9000/3.281)     #for Air density [kg/m3]
v = 125/1.944    # Air velocity [m/s]
# Procedure
N = 6 # Number of total stations for halft-wing
k = np.arange(1,N+1)
y_k = -(bw/2)*(1 - (2*k - 1)/(2*N))
c_k = Croot*(1 - ((2*(lambdaw - 1)/bw))*y_k)
beta_k = np.arcsin(-2 * y_k * htip / (bw * c_k))
a0_k = a0root - (2*y_k*( a0tip - a0root))/(bw)
aL_0_k = aL_0_root - (2*y_k*( aL_0_tip - aL_0_root))/(bw)
theta_k = np.arccos(-2*y_k/bw)
C = np.zeros((N,N))
for k_i in range(len(k)):
    for n in range(N):
        term1 = (4 * bw) / (a0_k[k_i] * c_k[k_i])
        term2 = (2 * (n + 1) - 1) / np.sin(theta_k[k_i])
        C[k_i, n] = (term1 + term2) * np.sin((2 * (n + 1) - 1) * theta_k[k_i])
        
C_inv = np.linalg.inv(C)

D_k = np.zeros((len(k), len(AoA)))
for ki in range(N):
    for AoAj in range(len(AoA)):
        D_k[ki,AoAj] = AoA[AoAj] - aL_0_k[ki] + beta_k[ki]
        
An = np.zeros((len(k), len(AoA)))
for AoAj in range(len(AoA)):
    An[:, AoAj] = np.dot(C_inv, D_k[:, AoAj])
CL_w = np.array([])
for AoAj in range(len(AoA)):
    CL_w = np.append(CL_w, An[0,AoAj] * np.pi * AR)
epsilon = 1e-10 

delta = []
for AoAi in range(len(AoA)):
    deltai = []
    for kd in range(1, N):
        if np.abs(An[0, AoAi]) < epsilon:
            term1 = 0  # O asigna un valor adecuado según el contexto
        else:
            term1 = (An[kd, AoAi] / An[0, AoAi])
        term1_square = term1**2
        deltai.append((kd + 1) * term1_square)
    deltai_sum = np.sum(deltai)
    delta.append(deltai_sum)
delta = np.array(delta)  
e = (1.0 + delta)**(-1)

CDi = []
for i in range(len(CL_w)):
    term1 = (CL_w[i])**2/(np.pi*e[i]*AR)
    CDi.append(term1)
CDi = np.array(CDi)
# -------------------------------------------------------------------------------------------------------------------
ruta_archivo = r"C:\Users\Lenovo\Documents\Universidad\2024-2\Aerodinámica\airfoilaerodynamicdataLLT.xlsx" #.------------DESCARGAR EL EXCEL Y CAMBIAR LA RUTA ESPECIFICADA AQUÍ DEPENDIENDO TU PC ------
#--------------------------------------------------------------------------------------------------------------------
# Luego, continúa leyendo el archivo de Excel
try:
    df = pd.read_excel(ruta_archivo, sheet_name='Hoja1')
except FileNotFoundError:
    print(f"No se puede encontrar el archivo de Excel en la ruta: {ruta_archivo}")
    sys.exit()
except PermissionError:
    print(f"No se tiene permiso para leer el archivo de Excel en la ruta: {ruta_archivo}")
    sys.exit()
except ValueError:
    print(f"La hoja 'Hoja1' no se encuentra en el archivo.")
    sys.exit()

# Extraer la columna deseada
cd_airfoil = df['cd_airfoil']
CD_w = np.add(CDi, cd_airfoil)
directorio_actual = os.path.dirname(os.path.abspath(__file__))

# Ruta relativa al archivo de configuración
ruta_config = os.path.join(directorio_actual, 'config.txt')

# Leer la ruta del archivo desde el archivo de configuración
with open(ruta_config, 'r') as file:
    nombre_archivo_excel = file.readline().strip()

# Ruta completa al archivo de Excel
ruta_excel = os.path.join(directorio_actual, nombre_archivo_excel)

# Guardar los resultados en el mismo archivo Excel
with pd.ExcelWriter(ruta_excel, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
    # Crear un DataFrame con ambas columnas
    df_resultados = pd.DataFrame({
        'Angle of Attack (deg)': AoA * 180 / np.pi,
        'CL_w': CL_w,
        'CD_w': CD_w
    })
    
    # Escribir los resultados en una sola hoja
    df_resultados.to_excel(writer, sheet_name='Resultados_CL_CD', index=False)

print("Resultados guardados en una sola hoja del archivo de Excel.")

# Convertir AoA de radianes a grados para mostrar en el eje X
AoA_deg = AoA * 180 / np.pi

# Ajustar el tamaño de la figura
plt.figure(figsize=(10, 8))  # Aumenta el tamaño de la figura

# Gráfico de CL_w vs Angle of Attack
plt.subplot(2, 1, 1)
plt.plot(AoA_deg, CL_w, label="CL_w", color='b')
plt.title('CL_w vs Angle of Attack')
plt.xlabel('Angle of Attack (deg)')
plt.ylabel('CL_w')

# Ajustar los ticks del eje X (ángulos de ataque) para que sean más congruentes
plt.xticks(np.linspace(min(AoA_deg), max(AoA_deg), 10))  # 10 valores distribuidos uniformemente
plt.yticks(np.linspace(min(CL_w), max(CL_w), 10))  # 10 valores distribuidos para CL_w

plt.grid(True)

# Gráfico de CD_w vs Angle of Attack
plt.subplot(2, 1, 2)
plt.plot(AoA_deg, CD_w, label="CD_w", color='r')
plt.title('CD_w vs Angle of Attack')
plt.xlabel('Angle of Attack (deg)')
plt.ylabel('CD_w')

# Ajustar los ticks del eje X e Y para que sean más congruentes
plt.xticks(np.linspace(min(AoA_deg), max(AoA_deg), 10))  # 10 valores distribuidos uniformemente
plt.yticks(np.linspace(min(CD_w), max(CD_w), 10))  # 10 valores distribuidos para CD_w

plt.grid(True)

# Ajustar el espaciado para evitar superposición
plt.tight_layout(pad=3.0)

# Guardar el gráfico como imagen
plt.savefig('resultados_graficos.png')

# Cargar el archivo de Excel existente
book = load_workbook(ruta_excel)

# Seleccionar o crear la hoja donde se guardarán los gráficos
if 'Resultados_CL_CD' in book.sheetnames:
    sheet = book['Resultados_CL_CD']
else:
    sheet = book.create_sheet('Resultados_CL_CD')

# Insertar la imagen en la hoja de Excel
img = Image('resultados_graficos.png')
img.width, img.height = 800, 600  # Ajustar tamaño de la imagen si es necesario
sheet.add_image(img, 'E5')  # Posición para colocar la imagen

# Guardar el archivo de Excel con los gráficos
book.save(ruta_excel)
book.close()

print("Gráficas generadas y guardadas en la hoja de Excel.")
print(e)